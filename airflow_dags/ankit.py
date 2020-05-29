import airflow
from airflow import DAG
from datetime import datetime
from datetime import datetime
from airflow.models import Variable
from airflow.operators.python_operator import BranchPythonOperator
from airflow.operators.python_operator import PythonOperator
import os
#import config
import pandas as pd
import openpyxl as x1 #Python library for editing excel files

args = {
    'owner': 'ankit',
    'depends_on_past': False,
    'email_on_retry': False,
    'start_date': datetime.utcnow(),
    'retries': 1,
    'retry_delay': timedelta(minutes=30)
    
}
dag = DAG(dag_id='test', default_args=args, schedule_interval='@hourly')

def check_time_modified():
    
	mod_time = Variable.get('modified_time')
	filename = '/Users/Ankit/Desktop/songs'
	stamp = os.stat(filename).st_mtime
	if stamp != mod_time:
			self._cached_stamp = stamp
			Variable.set('modified_time', stamp)
            return 'make_list_of_files_in_directory'
    else 
            return 'print_no_change_in_directory'

def print_no_change_in_directory():
    logging.info("Directory has not been changed")




def make_list_of_files_in_directory():
    arr=os.listdir('/Users/Ankit/Desktop/songs')
    filename ="/Users/Ankit/Desktop/current.xlsx"
    
    # wb = xl.load_workbook(filename)
    # sheet  = wb.worksheets[0]
    # for i in range of len(arr)
    #     data = sheet.cell(row = i, column = 1)
    #     data.value = arr[0]

def find_new_files_added():

    # df_A = pd.read_excel("/Users/Ankit/Desktop/current.xlsx", dtype=str)
    # df_B = pd.read_excel("/Users/Ankit/Desktop/previous.xlsx", dtype=str) 
    # df_new = df_A.merge(df_B, on = 'ID',how='outer',indicator=True)
    # df_common = df_new[df_new['_merge'] == 'both']
    # df_A = df_A[(~df_A.ID.isin(df_common.ID))]        

def change_previous_to_current():
    # wb1 = xl.load_workbook("/Users/Ankit/Desktop/current.xlsx")
    # ws1 = wb1.worksheets[0]

    # wb2 = xl.load_workbook("/Users/Ankit/Desktop/previous.xlsx") 
    # ws2 = wb2.active 
    
    # mr = ws1.max_row 
    # mc = ws1.max_column

    # for i in range (1, mr + 1):
    #     for j in range (1, mc + 1):
    #         c = ws1.cell(row = i, column = j)
    #         ws2.cell(row = i, column = j).value = c.value

    # wb2.save(str("/Users/Ankit/Desktop/previous.xlsx"))   



t1 = BranchPythonOperator(
  task_id='task1', 
  python_callable=check_time_modified, 
  dag=dag)

t2 = PythonOperator(
  task_id='task2', 
  python_callable=print_no_change_in_directory, 
  dag=dag)

t3 = PythonOperator(
  task_id='task3', 
  python_callable=make_list_of_files_in_directory, 
  dag=dag)

t4 = PythonOperator(
  task_id='task4', 
  python_callable=find_new_files_added, 
  dag=dag)

t5 = PythonOperator(
  task_id='task5', 
  python_callable=change_previous_to_current, 
  dag=dag)

t1 >> [t2, t3]
t3 >> t4 >> t5


