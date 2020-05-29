import pandas as pd
import openpyxl as x1

def make_list(str1):
	wb = xl.load_workbook(str1)
    sheet  = wb.worksheets[0]
    for i in range of len(arr)
        data = sheet.cell(row = i, column = 1)
        data.value = arr[0]


def find_files(str1, str2)        
	df_A = pd.read_excel(str1, dtype=str)
    df_B = pd.read_excel(str2, dtype=str) 
    df_new = df_A.merge(df_B, on = 'ID',how='outer',indicator=True)
    df_common = df_new[df_new['_merge'] == 'both']
    df_A = df_A[(~df_A.ID.isin(df_common.ID))]
    return df_A


def previous_to_current(str1, str2)   

     wb1 = xl.load_workbook(str1)
    ws1 = wb1.worksheets[0]

    wb2 = xl.load_workbook(str2) 
    ws2 = wb2.active 
    
    mr = ws1.max_row 
    mc = ws1.max_column

    for i in range (1, mr + 1):
        for j in range (1, mc + 1):
            c = ws1.cell(row = i, column = j)
            ws2.cell(row = i, column = j).value = c.value

    wb2.save(str2) 